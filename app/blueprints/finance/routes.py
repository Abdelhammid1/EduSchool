from datetime import datetime, date, timedelta
from decimal import Decimal
from io import BytesIO

from flask import (
    abort, current_app, flash, redirect, render_template, request, Response, url_for,
)
from flask_login import current_user, login_required
from sqlalchemy import func

from . import bp
from ..utils import require_permission
from ...extensions import db
from ...models import (
    Account, AcademicYear, Enrollment, Expense, FeeType, Grade,
    Installment, Invoice, InvoiceLine, JournalEntry, JournalLine,
    Payment, Section, Student, Vendor,
)
from ...services.accounting import post_journal
from ...services.notifications import send_notification


def _sid():
    return current_user.school_id


def _get(model, oid):
    obj = model.query.filter_by(id=oid, school_id=_sid()).first()
    if not obj:
        abort(404)
    return obj


def _ar_account() -> Account:
    return Account.query.filter_by(school_id=_sid(), code="1300").first()


def _default_cash_account() -> Account:
    return Account.query.filter_by(school_id=_sid(), code="1100").first()


# ---------- T-8.1 Chart of accounts ----------

@bp.route("/accounts")
@login_required
@require_permission("finance", "view")
def accounts():
    items = (
        Account.query.filter_by(school_id=_sid())
        .order_by(Account.code).all()
    )
    return render_template("finance/accounts.html", accounts=items)


@bp.route("/accounts/new", methods=["GET", "POST"])
@login_required
@require_permission("finance", "edit")
def account_new():
    parents = Account.query.filter_by(school_id=_sid()).order_by(Account.code).all()
    if request.method == "POST":
        a = Account(
            school_id=_sid(),
            code=request.form["code"].strip(),
            name=request.form["name"].strip(),
            type=request.form["type"],
            parent_id=int(request.form["parent_id"]) if request.form.get("parent_id") else None,
        )
        db.session.add(a)
        db.session.commit()
        flash(f"تم إضافة الحساب {a.code} — {a.name}.", "success")
        return redirect(url_for("finance.accounts"))
    return render_template("finance/account_form.html", parents=parents)


# ---------- T-8.1 Journal view ----------

@bp.route("/journal")
@login_required
@require_permission("finance", "view")
def journal():
    end = _parse_date(request.args.get("end")) or date.today()
    start = _parse_date(request.args.get("start")) or (end - timedelta(days=365))
    account_id = request.args.get("account_id", type=int)
    accounts = Account.query.filter_by(school_id=_sid()).order_by(Account.code).all()

    q = JournalEntry.query.filter(
        JournalEntry.school_id == _sid(),
        JournalEntry.entry_date >= start,
        JournalEntry.entry_date <= end,
    )
    if account_id:
        q = q.join(JournalLine).filter(JournalLine.account_id == account_id)
    entries = q.order_by(JournalEntry.entry_date.desc(), JournalEntry.id.desc()).limit(200).all()
    return render_template(
        "finance/journal.html",
        entries=entries, accounts=accounts,
        start=start, end=end, account_id=account_id,
    )


# ---------- T-8.2 Fee types ----------

@bp.route("/fee-types")
@login_required
@require_permission("finance", "view")
def fee_types():
    items = FeeType.query.filter_by(school_id=_sid()).order_by(FeeType.name).all()
    return render_template("finance/fee_types.html", fee_types=items)


@bp.route("/fee-types/new", methods=["GET", "POST"])
@login_required
@require_permission("finance", "edit")
def fee_type_new():
    revenues = (
        Account.query.filter_by(school_id=_sid(), type="revenue")
        .order_by(Account.code).all()
    )
    if request.method == "POST":
        f = FeeType(
            school_id=_sid(),
            name=request.form["name"].strip(),
            default_amount=Decimal(request.form.get("default_amount") or "0"),
            installable=bool(request.form.get("installable")),
            revenue_account_id=int(request.form["revenue_account_id"]),
        )
        db.session.add(f)
        db.session.commit()
        flash("تم إضافة نوع الرسم.", "success")
        return redirect(url_for("finance.fee_types"))
    return render_template("finance/fee_type_form.html", revenues=revenues)


# ---------- T-8.3 / T-8.4 / T-8.5 Invoices ----------

@bp.route("/invoices")
@login_required
@require_permission("finance", "view")
def invoices_list():
    year = AcademicYear.query.filter_by(school_id=_sid(), status="active").first()
    q = (
        Invoice.query.filter_by(school_id=_sid())
        .join(Enrollment).join(Student)
        .order_by(Invoice.issue_date.desc())
    )
    if year:
        q = q.filter(Enrollment.year_id == year.id)
    invoices = q.limit(500).all()
    return render_template("finance/invoices_list.html", invoices=invoices, year=year)


@bp.route("/invoices/new", methods=["GET", "POST"])
@login_required
@require_permission("finance", "edit")
def invoice_new():
    year = AcademicYear.query.filter_by(school_id=_sid(), status="active").first()
    if not year:
        flash("لا توجد سنة دراسية نشطة.", "warning")
        return redirect(url_for("finance.invoices_list"))

    enrollments = (
        Enrollment.query.filter_by(school_id=_sid(), year_id=year.id, status="active")
        .join(Student).order_by(Student.full_name).all()
    )
    fee_types = FeeType.query.filter_by(school_id=_sid(), is_active=True).order_by(FeeType.name).all()
    ar = _ar_account()
    if not ar:
        flash("لم يُهيّأ دليل الحسابات. أنشئ الحسابات الأساسية أولاً.", "danger")
        return redirect(url_for("finance.accounts"))

    if request.method == "POST":
        enrollment_id = int(request.form["enrollment_id"])
        issue = _parse_date(request.form.get("issue_date")) or date.today()
        due = _parse_date(request.form.get("due_date")) or (issue + timedelta(days=30))
        installments_count = int(request.form.get("installments_count") or "1")

        fee_ids = request.form.getlist("fee_type_id", type=int)
        amounts = request.form.getlist("amount")
        if not fee_ids:
            flash("أضف نوع رسم واحد على الأقل.", "danger")
            return redirect(url_for("finance.invoice_new"))

        # Invoice number
        n = Invoice.query.filter_by(school_id=_sid()).count() + 1
        number = f"INV-{year.name}-{n:05d}"

        inv = Invoice(
            school_id=_sid(),
            enrollment_id=enrollment_id,
            number=number,
            issue_date=issue,
            due_date=due,
            status="sent",
        )
        db.session.add(inv)
        db.session.flush()

        total = Decimal(0)
        rev_lines = {}
        for fid, amt_raw in zip(fee_ids, amounts):
            amt = Decimal(amt_raw or "0")
            if amt <= 0:
                continue
            ft = FeeType.query.filter_by(school_id=_sid(), id=fid).first()
            if not ft:
                continue
            line = InvoiceLine(
                invoice_id=inv.id, fee_type_id=ft.id,
                description=ft.name, amount=amt,
            )
            db.session.add(line)
            total += amt
            rev_lines[ft.revenue_account_id] = rev_lines.get(ft.revenue_account_id, Decimal(0)) + amt

        inv.total_amount = total

        # Split installments
        if installments_count < 1:
            installments_count = 1
        per = (total / installments_count).quantize(Decimal("0.01"))
        accum = Decimal(0)
        for i in range(installments_count):
            d = due if installments_count == 1 else due + timedelta(days=30 * i)
            amt = per if i < installments_count - 1 else total - accum
            db.session.add(Installment(
                invoice_id=inv.id, due_date=d, amount=amt,
            ))
            accum += amt

        # Auto-journal: DR AR, CR revenue accounts
        journal_lines = [(ar.id, total, Decimal(0), "ذمم الطالب")]
        for rev_id, amt in rev_lines.items():
            journal_lines.append((rev_id, Decimal(0), amt, "إيراد رسوم"))
        post_journal(
            school_id=_sid(),
            entry_date=issue,
            description=f"فاتورة {number}",
            reference=number,
            lines=journal_lines,
            related_kind="invoice",
            related_id=inv.id,
        )
        db.session.commit()

        # T-8.5: notify parent on issue
        e = inv.enrollment
        phone = (e.student.parent_phone or "").strip()
        if phone:
            send_notification(
                school_id=_sid(),
                kind="invoice_issued",
                payload={
                    "student": e.student.full_name,
                    "invoice_number": number,
                    "amount": float(total),
                    "due_date": due.isoformat(),
                    "message": (
                        f"إشعار فاتورة: صدرت فاتورة برقم {number} للطالب "
                        f"{e.student.full_name} بمبلغ {total} مستحقة بتاريخ {due.isoformat()}."
                    ),
                },
                target_phone=phone,
                related_kind="invoice", related_id=inv.id,
            )

        flash(f"تم إنشاء الفاتورة {number} وقيدها محاسبيًا.", "success")
        return redirect(url_for("finance.invoice_detail", invoice_id=inv.id))

    return render_template(
        "finance/invoice_form.html", year=year, enrollments=enrollments, fee_types=fee_types,
    )


@bp.route("/invoices/<int:invoice_id>")
@login_required
@require_permission("finance", "view")
def invoice_detail(invoice_id):
    inv = _get(Invoice, invoice_id)
    cash_accounts = Account.query.filter_by(school_id=_sid(), type="asset").order_by(Account.code).all()
    return render_template("finance/invoice_detail.html", inv=inv, cash_accounts=cash_accounts)


@bp.route("/invoices/<int:invoice_id>/print")
@login_required
@require_permission("finance", "view")
def invoice_print(invoice_id):
    """Sprint 9 TC-8.3.1 — print-optimized invoice view.

    ?download=1 pipes through WeasyPrint to return a PDF.
    Default: renders HTML with @media print styles + auto window.print().
    """
    inv = _get(Invoice, invoice_id)
    download = request.args.get("download") == "1"
    if download:
        html = render_template("finance/invoice_print.html", inv=inv, download=True)
        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()
        except Exception as e:  # noqa: BLE001
            current_app.logger.exception("WeasyPrint invoice PDF failed: %s", e)
            # Fallback: return the printable HTML so the browser can Save-as-PDF
            return Response(html, mimetype="text/html")
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="invoice_{inv.number}.pdf"',
            },
        )
    return render_template("finance/invoice_print.html", inv=inv, download=False)


@bp.route("/invoices/<int:invoice_id>/pay", methods=["POST"])
@login_required
@require_permission("finance", "edit")
def invoice_pay(invoice_id):
    inv = _get(Invoice, invoice_id)
    amount = Decimal(request.form["amount"])
    if amount <= 0:
        flash("المبلغ غير صالح.", "danger")
        return redirect(url_for("finance.invoice_detail", invoice_id=inv.id))

    is_refund = request.form.get("is_refund") == "1"
    method = request.form.get("method", "cash")
    cash_account_id = int(request.form["cash_account_id"])
    cash = _get(Account, cash_account_id)
    ar = _ar_account()
    if not ar:
        abort(400)

    pay_date = _parse_date(request.form.get("payment_date")) or date.today()

    if is_refund:
        if amount > Decimal(str(inv.paid_amount)):
            flash("لا يمكن استرداد مبلغ أكبر من المدفوع.", "danger")
            return redirect(url_for("finance.invoice_detail", invoice_id=inv.id))
        inv.paid_amount = Decimal(str(inv.paid_amount)) - amount
        # Reverse installment distribution LIFO (latest-paid first) so
        # installment.paid_amount tracks invoice.paid_amount exactly
        refund_remaining = amount
        for inst in reversed(list(inv.installments)):
            if refund_remaining <= 0:
                break
            paid = Decimal(str(inst.paid_amount))
            if paid <= 0:
                continue
            take = min(paid, refund_remaining)
            inst.paid_amount = paid - take
            inst.status = "paid" if Decimal(str(inst.remaining)) <= 0 else "pending"
            refund_remaining -= take
        je = post_journal(
            school_id=_sid(),
            entry_date=pay_date,
            description=f"استرداد على الفاتورة {inv.number}",
            reference=inv.number,
            lines=[
                (ar.id, amount, Decimal(0), "إعادة الذمة"),
                (cash.id, Decimal(0), amount, "خروج نقدية"),
            ],
            related_kind="payment", related_id=inv.id,
        )
        flash(f"تم استرداد {amount} من الفاتورة {inv.number}.", "success")
    else:
        if amount > Decimal(str(inv.remaining)):
            flash("المبلغ أكبر من المتبقّي.", "danger")
            return redirect(url_for("finance.invoice_detail", invoice_id=inv.id))
        inv.paid_amount = Decimal(str(inv.paid_amount)) + amount
        je = post_journal(
            school_id=_sid(),
            entry_date=pay_date,
            description=f"دفعة على الفاتورة {inv.number}",
            reference=inv.number,
            lines=[
                (cash.id, amount, Decimal(0), "تحصيل نقدي"),
                (ar.id, Decimal(0), amount, "تخفيض ذمم"),
            ],
            related_kind="payment", related_id=inv.id,
        )
        # Distribute payment across installments
        remaining = amount
        for inst in inv.installments:
            if remaining <= 0: break
            r = Decimal(str(inst.remaining))
            if r <= 0: continue
            take = min(r, remaining)
            inst.paid_amount = Decimal(str(inst.paid_amount)) + take
            inst.status = "paid" if inst.remaining <= 0 else "pending"
            remaining -= take

        flash(f"تم تسجيل دفعة {amount} على الفاتورة {inv.number}.", "success")

    # Recompute status
    if inv.paid_amount <= 0:
        inv.status = "refunded" if is_refund else "sent"
    elif inv.paid_amount >= inv.total_amount:
        inv.status = "paid"
    else:
        inv.status = "partial"

    payment = Payment(
        school_id=_sid(),
        invoice_id=inv.id,
        payment_date=pay_date,
        amount=amount,
        method=method,
        cash_account_id=cash.id,
        is_refund=is_refund,
        reference=(request.form.get("reference") or "").strip() or None,
        notes=(request.form.get("notes") or "").strip() or None,
        journal_entry_id=je.id,
    )
    db.session.add(payment)
    db.session.commit()

    # T-8.5: notify on payment
    phone = (inv.enrollment.student.parent_phone or "").strip()
    if phone:
        kind = "refund" if is_refund else "payment"
        send_notification(
            school_id=_sid(),
            kind=kind,
            payload={
                "student": inv.enrollment.student.full_name,
                "invoice_number": inv.number,
                "amount": float(amount),
                "remaining": float(inv.remaining),
                "message": (
                    f"إشعار {'استرداد' if is_refund else 'دفع'}: "
                    f"تم تسجيل {amount} على الفاتورة {inv.number}. "
                    f"المتبقّي: {inv.remaining}."
                ),
            },
            target_phone=phone,
            related_kind="payment", related_id=payment.id,
        )

    return redirect(url_for("finance.invoice_detail", invoice_id=inv.id))


# ---------- T-9.1 Vendors + Expenses ----------

@bp.route("/vendors")
@login_required
@require_permission("expenses", "view")
def vendors_list():
    items = Vendor.query.filter_by(school_id=_sid()).order_by(Vendor.name).all()
    return render_template("finance/vendors_list.html", vendors=items)


@bp.route("/vendors/new", methods=["GET", "POST"])
@login_required
@require_permission("expenses", "edit")
def vendor_new():
    if request.method == "POST":
        v = Vendor(
            school_id=_sid(),
            name=request.form["name"].strip(),
            phone=(request.form.get("phone") or "").strip() or None,
            email=(request.form.get("email") or "").strip() or None,
            address=(request.form.get("address") or "").strip() or None,
        )
        db.session.add(v)
        db.session.commit()
        flash("تم إضافة المورد.", "success")
        return redirect(url_for("finance.vendors_list"))
    return render_template("finance/vendor_form.html")


@bp.route("/expenses")
@login_required
@require_permission("expenses", "view")
def expenses_list():
    end = _parse_date(request.args.get("end")) or date.today()
    start = _parse_date(request.args.get("start")) or (end - timedelta(days=90))
    items = (
        Expense.query.filter_by(school_id=_sid())
        .filter(Expense.date >= start, Expense.date <= end)
        .order_by(Expense.date.desc()).all()
    )
    total = sum((Decimal(str(e.amount)) for e in items), Decimal(0))
    return render_template(
        "finance/expenses_list.html",
        expenses=items, total=total, start=start, end=end,
    )


@bp.route("/expenses/new", methods=["GET", "POST"])
@login_required
@require_permission("expenses", "edit")
def expense_new():
    vendors = Vendor.query.filter_by(school_id=_sid(), is_active=True).order_by(Vendor.name).all()
    exp_accounts = Account.query.filter_by(school_id=_sid(), type="expense").order_by(Account.code).all()
    cash_accounts = Account.query.filter_by(school_id=_sid(), type="asset").order_by(Account.code).all()
    if request.method == "POST":
        amount = Decimal(request.form["amount"])
        d = _parse_date(request.form.get("date")) or date.today()
        ex_account = _get(Account, int(request.form["expense_account_id"]))
        cash = _get(Account, int(request.form["cash_account_id"]))
        je = post_journal(
            school_id=_sid(),
            entry_date=d,
            description=f"مصروف: {request.form['description'].strip()}",
            reference=(request.form.get("reference") or "").strip() or None,
            lines=[
                (ex_account.id, amount, Decimal(0), "مصروف"),
                (cash.id, Decimal(0), amount, "خروج نقدية"),
            ],
            related_kind="expense", related_id=None,
        )
        e = Expense(
            school_id=_sid(),
            vendor_id=int(request.form["vendor_id"]) if request.form.get("vendor_id") else None,
            expense_account_id=ex_account.id,
            cash_account_id=cash.id,
            date=d, amount=amount,
            description=request.form["description"].strip(),
            reference=(request.form.get("reference") or "").strip() or None,
            journal_entry_id=je.id,
        )
        db.session.add(e)
        db.session.commit()
        flash(f"تم تسجيل المصروف ({amount}) وقيده محاسبيًا.", "success")
        return redirect(url_for("finance.expenses_list"))
    return render_template(
        "finance/expense_form.html",
        vendors=vendors, exp_accounts=exp_accounts, cash_accounts=cash_accounts,
    )


# ---------- T-9.2 Reports ----------

def _report_data(start, end):
    """Sprint 9 TC-9.2.2 — shared computation for HTML view + PDF/Excel export."""
    sums = (
        db.session.query(
            Account.type, Account.id, Account.code, Account.name,
            func.coalesce(func.sum(JournalLine.debit), 0).label("d"),
            func.coalesce(func.sum(JournalLine.credit), 0).label("c"),
        )
        .join(JournalLine, JournalLine.account_id == Account.id)
        .join(JournalEntry, JournalEntry.id == JournalLine.entry_id)
        .filter(Account.school_id == _sid())
        .filter(JournalEntry.entry_date >= start, JournalEntry.entry_date <= end)
        .group_by(Account.id, Account.type, Account.code, Account.name)
        .order_by(Account.code).all()
    )

    def bal(t, d, c):
        d, c = float(d), float(c)
        return d - c if t in ("asset", "expense") else c - d

    grouped = {"asset": [], "liability": [], "equity": [], "revenue": [], "expense": []}
    totals = {k: 0.0 for k in grouped}
    for t, aid, code, name, d, c in sums:
        b = bal(t, d, c)
        grouped.setdefault(t, []).append({"id": aid, "code": code, "name": name, "balance": b})
        totals[t] = totals.get(t, 0.0) + b

    net_income = totals["revenue"] - totals["expense"]
    return grouped, totals, net_income


@bp.route("/reports")
@login_required
@require_permission("finance", "view")
def reports():
    end = _parse_date(request.args.get("end")) or date.today()
    start = _parse_date(request.args.get("start")) or (end - timedelta(days=365))
    grouped, totals, net_income = _report_data(start, end)
    return render_template(
        "finance/reports.html",
        start=start, end=end, grouped=grouped, totals=totals, net_income=net_income,
    )


@bp.route("/reports/export")
@login_required
@require_permission("finance", "view")
def reports_export():
    """Sprint 9 TC-9.2.2 — export the financial report as PDF or Excel."""
    fmt = (request.args.get("format") or "pdf").lower()
    end = _parse_date(request.args.get("end")) or date.today()
    start = _parse_date(request.args.get("start")) or (end - timedelta(days=365))
    grouped, totals, net_income = _report_data(start, end)

    label_ar = {
        "asset": "الأصول", "liability": "الالتزامات", "equity": "حقوق الملكية",
        "revenue": "الإيرادات", "expense": "المصروفات",
    }

    if fmt == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "التقرير المالي"
        ws.sheet_view.rightToLeft = True

        bold_navy = Font(bold=True, color="FFFFFF")
        navy_fill = PatternFill("solid", fgColor="001556")
        gold_fill = PatternFill("solid", fgColor="D4AF37")
        header_font = Font(bold=True, size=13, color="001556")

        ws["A1"] = "التقرير المالي"
        ws["A1"].font = Font(bold=True, size=16, color="001556")
        ws.merge_cells("A1:D1")
        ws["A2"] = f"من {start.isoformat()} إلى {end.isoformat()}"
        ws["A2"].alignment = Alignment(horizontal="right")

        r = 4
        for section_key in ("asset", "liability", "equity", "revenue", "expense"):
            ws.cell(row=r, column=1, value=label_ar[section_key]).font = header_font
            r += 1
            headers = ["الرمز", "الاسم", "الرصيد"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=c, value=h)
                cell.font = bold_navy
                cell.fill = navy_fill
                cell.alignment = Alignment(horizontal="center")
            r += 1
            for entry in grouped[section_key]:
                ws.cell(row=r, column=1, value=entry["code"])
                ws.cell(row=r, column=2, value=entry["name"])
                ws.cell(row=r, column=3, value=round(entry["balance"], 2))
                r += 1
            tot_cell = ws.cell(row=r, column=2, value="الإجمالي")
            tot_cell.font = Font(bold=True)
            total_val_cell = ws.cell(row=r, column=3, value=round(totals[section_key], 2))
            total_val_cell.font = Font(bold=True)
            total_val_cell.fill = gold_fill
            r += 2

        ws.cell(row=r, column=1, value="صافي الدخل").font = header_font
        ws.cell(row=r, column=3, value=round(net_income, 2)).font = Font(bold=True, size=14)
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="financial_report_{start}_{end}.xlsx"'
                ),
            },
        )

    # Default: PDF via WeasyPrint (renders the print-optimized template).
    #
    # Sprint 10 hotfix — WeasyPrint has native deps (Pango, Cairo, GDK-PixBuf,
    # HarfBuzz). On production servers those must be installed with the OS
    # package manager (apt install libpango-1.0-0 libpangoft2-1.0-0 …).
    # If the deps are missing, the `import weasyprint` line raises OSError.
    # We catch it and fall back to a printable HTML page + a clear Arabic
    # error message so the user isn't stuck on a bare 500.
    html = render_template(
        "finance/reports_pdf.html",
        start=start, end=end,
        grouped=grouped, totals=totals, net_income=net_income,
        label_ar=label_ar,
    )
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()
    except Exception as e:  # noqa: BLE001
        current_app.logger.exception("WeasyPrint PDF export failed: %s", e)
        # Fallback: return the printable HTML with a banner asking the user
        # to use browser print (⌘/Ctrl+P → Save as PDF).
        fallback = (
            "<!doctype html><html lang='ar' dir='rtl'><head><meta charset='utf-8'>"
            "<title>تقرير مالي — عرض للطباعة</title></head><body style='font-family:sans-serif'>"
            "<div style='background:#fef3c7;border:1px solid #d4af37;padding:12px 16px;"
            "margin-bottom:16px;border-radius:6px;color:#1a2d6c'>"
            "<strong>ملاحظة:</strong> تعذّر توليد ملف PDF على السيرفر "
            "(قد تكون مكتبات WeasyPrint الأساسية غير مثبَّتة). "
            "استخدم أمر <strong>طباعة</strong> في المتصفح "
            "(⌘+P أو Ctrl+P) واختر <em>حفظ كملف PDF</em>."
            "</div>"
            + html.split("<body>", 1)[-1].split("</body>", 1)[0]
            + "</body></html>"
        )
        return Response(fallback, mimetype="text/html")

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="financial_report_{start}_{end}.pdf"'
            ),
        },
    )


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()
